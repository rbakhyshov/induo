# -*- coding: utf-8 -*-

import telebot
from settings import *

import requests
from bs4 import BeautifulSoup, SoupStrainer
import openpyxl as xl
import time, datetime
import re
from selenium import webdriver

import MPAsync_lib as lib

MainDomain = 'https://induo.store'

TAG = '?utm_content=catalog'

items_data = []
items_pages = []
capsule_pages = []

linksAllowed = ['/', '#',
                'tel:+79166819944', 'https://wa.clck.bar/79166819944', 'mailto:info@induo.store',
                'https://wa.me/79166819944', 'https://t.me/InDuoStore', 'https://alphy.ru/']

linksIgnore = ['#',
                'tel:+79166819944', 'https://wa.clck.bar/79166819944', 'mailto:info@induo.store',
                'https://wa.me/79166819944', 'https://t.me/InDuoStore', 'https://alphy.ru/',
                '/bags', '/shoes', '/clothes', '/accessories', '/musthave', '/capsules', '/brands',
                '/fashion_concierge', '/ask_a_stylist', '/info']

Capsules = {'1': 4,
            '2': 5,
            '3': 7,
            '4': 4,
            '5': 6,
            '6': 5,
            '7': 3,
            '8': 2,
            '9': 4,
            '10': 5,
            '11': 5,
            '12': 4,
            '13': 3,
            '14':5}

def save_xls (file_name, items_data, capsule_pages):

    wb = xl.load_workbook(f'../output/induo_temp.xlsx')
    ws = wb.worksheets[0]
    xlrow = 2

    for item in items_data:
        ws.cell(xlrow, 1, ', '.join(item["warn"].keys()))
        ws.cell(xlrow, 2, item["itemBrandD"])
        ws.cell(xlrow, 3, item["itemNameD"])
        try:
            ws.cell(xlrow, 4, int(item["itemPrice"]))       #пробуем записать числом, чтобы корректно отобразилось в excel
        except:
            ws.cell(xlrow, 4, item["itemPrice"])
        ws.cell(xlrow, 5, item["itemSizesStr"].strip())
        ws.cell(xlrow, 6, item["itemDescD"].strip())

        ws.cell(xlrow, 7).value = '\n'.join(item["otherColorsHref"]).strip()
        ws.cell(xlrow, 8).value = '\n'.join(item["missedLinks"]).strip()

        #ws.cell(xlrow, 9).value = '\n'.join(item["capsuleLinks"]).strip()
        #ws.cell(xlrow, 10).value = '\n'.join(item["missedCapsuleLinks"]).strip()
        ws.cell(xlrow, 9).value = '\n'.join(item["foreignLinks"]).strip()

        ws.cell(xlrow, 10, item["pageTitle"])
        ws.cell(xlrow, 11, item["itemURL"])
        ws.cell(xlrow, 12, item["chapter"])

        ws.cell(xlrow, 13, item["itemBrandD"])
        ws.cell(xlrow, 14, item["itemBrandM"])
        ws.cell(xlrow, 15, item["catBrand"])

        ws.cell(xlrow, 16, item["itemNameD"])
        ws.cell(xlrow, 17, item["itemNameM"])
        ws.cell(xlrow, 18, item["catName"])

        ws.cell(xlrow, 19, item["itemDescD"].strip())
        ws.cell(xlrow, 20, item["itemDescM"].strip())

        xlrow = xlrow + 1

    ws = wb.worksheets[1]
    xlrow = 2
    for capsule in capsule_pages:
        ws.cell(xlrow, 1, ', '.join(capsule["warn"].keys()))
        ws.cell(xlrow, 2, capsule["code"])
        ws.cell(xlrow, 3, capsule["URL"])
        ws.cell(xlrow, 4).value = '\n'.join(capsule["items"]).strip()
        ws.cell(xlrow, 5).value = '\n'.join(capsule["missedItems"]).strip()
        ws.cell(xlrow, 6, capsule["cnt"])
        xlrow = xlrow + 1

    cur_time = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M")
    file_name = f'../output/{file_name}_{cur_time}.xlsx'

    wb.save(file_name)
    return file_name

def get_page_data(data, browser, is_debug = False):

    url = data["URL"]
    #print(url)
    chapter = data["Chapter"][len(MainDomain)+1:] #выкусываем chapter
    catName = data["Name"]
    catBrand = data["Brand"]

    browser.get(url)
    time.sleep(0.2)
    response = browser.page_source

    itemBlock = BeautifulSoup(response, "lxml")

    #response = requests.get(url=url, headers=lib.HEADERS)
    #itemBlock = BeautifulSoup(response.text, "lxml")

    pageTitle = itemBlock.find('title').text.strip()

    foreignLinks = []
    for link in itemBlock.find_all('a', href=True):
        flag = False
        for l in linksAllowed:
            if link['href'][0:len(l)] == l:
                flag = True
                break
        if not flag:
            foreignLinks.append(link['href'])

    try:
        itemBrandM = itemBlock.find('div', class_='t762__title_small t-descr t-descr_xxs js-product-sku js-store-prod-sku').text.replace('#', '').strip()
    except:
        itemBrandM = ''

    try:
        itemBrandD = itemBlock.find('div', class_='t760__title_small t-descr t-descr_xxs js-product-sku js-store-prod-sku').text.replace('#', '').strip()
    except:
        itemBrandD = ''


    try:
        itemNameM = itemBlock.find('div', class_='t762__title t-name t-name_xl js-product-name').text.strip()
    except:
        itemNameM = ''

    try:
        itemNameD = itemBlock.find('div', class_='t760__title t-name t-name_xl js-product-name').text.strip()
    except:
        itemNameD = ''

    try:
        itemPrice = itemBlock.find('div', field='price').text.strip()
        itemPrice = itemPrice.replace('$', '').replace(' ', '').replace('.', ',').strip()
    except:
        itemPrice = ''

    try:
        itemDescM = itemBlock.find('div', class_= 't762__descr t-descr t-descr_xxs js-store-prod-text').text#.strip() t762__descr t-descr t-descr_xxs
    except:
        itemDescM = ''

    try:
        itemDescD = itemBlock.find('div', class_='t760__descr t-descr t-descr_xxs js-store-prod-text').text  # .strip()
    except:
        itemDescD = ''

    try:
        otherColorsHref = []
        otherColors = itemBlock.find_all('div', class_=(re.compile('t396__elem tn-elem tn-elem')))
        for o in otherColors:
            try:
                h = o.find('a').get('href')
                if h != '/':
                    flag = False
                    for l in linksIgnore:
                        if (h[0:len(l)] == l) and (h != '/'):
                            flag = True
                            break
                    if not flag:
                        otherColorsHref.append(h)
            except:
                pass
    except:
        pass

    try:
        itemSizes = itemBlock.find('div', class_='t-product__option-variants').find_all('option')
        itemSizesStr = ''
        for size in itemSizes:
            itemSizesStr = itemSizesStr + size.text + ", "
        itemSizesStr = itemSizesStr[:-2]
    except:
        itemSizesStr = ''

    try:
        capsuleLinks =[]
        capsuleItems = itemBlock.find_all('div', class_='js-product t-store__card t-store__stretch-col t-store__stretch-col_33 t-align_left t-item')
        for c in capsuleItems:
            try:
                h = c.find('a').get('href')
                capsuleLinks.append(h)
            except:
                pass
    except:
        pass

    items_data.append(
        {
            "itemBrandM": itemBrandM,
            "itemBrandD": itemBrandD,
            "itemNameM": itemNameM,
            "itemNameD": itemNameD,
            "itemPrice": itemPrice,
            "itemSizesStr" :itemSizesStr,
            "itemDescM": itemDescM,
            "itemDescD": itemDescD,
            "otherColorsHref": otherColorsHref,
            "itemURL": url,
            "foreignLinks": foreignLinks,
            "chapter": chapter,
            "catName": catName,
            "catBrand": catBrand,
            "pageTitle": pageTitle,
            "capsuleLinks": capsuleLinks,
            "warn": [],
            "missedLinks": [],
            "missedCapsuleLinks": []
        }
    )

    if is_debug:
        print("debug get_page_data")
        print(f"pageTitle = {pageTitle}")
        print(f"url = {url}")
        print(f"itemBrandM = {itemBrandM}")
        print(f"itemBrandD = {itemBrandD}")
        print(f"itemNameM = {itemNameM}")
        print(f"itemNameD = {itemNameD}")
        print(f"itemPrice = {itemPrice}")
        print(f"itemSizesStr = {itemSizesStr}")
        print(f"itemDescM = {itemDescM}")
        print(f"itemDescD = {itemDescD}")
        print(f"otherColorsHref = {otherColorsHref}")
        print(f"foreignLinks = {foreignLinks}")
        print(f"capsuleLinks = {capsuleLinks}")
        print(f"chapter = {chapter}")

        print(f"------------------------")
        print(f"itemBlock = {itemBlock}")

def get_capsules ():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    browser = webdriver.Chrome(options=options)
    browser.get(MainDomain + '/capsules')
    time.sleep(1)
    response = browser.page_source
    soup = BeautifulSoup(response, "lxml")

    pageBlocks = soup.find_all('div', class_=(re.compile('t396__elem tn-elem tn-elem')))

    for pageBlock in pageBlocks:
        p = pageBlock.find('a', class_ = 'tn-atom')
        try:
            href = p.get('href')
            if href[0:10] == '/capsules/':
                capsuleCode = href.split('/')[-1]
                capsuleURL = MainDomain + href
                cnt = Capsules[capsuleCode]
                capsule_pages.append({"code": capsuleCode,
                                      "URL": capsuleURL,
                                      "cnt": cnt,
                                      "items":[],
                                      "warn" :[],
                                      "missedItems":[]})
        except:
            pass

    for page in capsule_pages:
        #print (page["URL"])
        browser.get(page["URL"])
        time.sleep(.3)
        response = browser.page_source
        soup = BeautifulSoup(response, "lxml")

        items = soup.find_all('div', class_= 'js-product t-store__card t-store__stretch-col t-store__stretch-col_25 t-align_left t-item')
        for item in items:
            try:
                itemURL = item.find('a').get('href')
                page["items"].append(itemURL)
            except:
                pass
    return capsule_pages
    #print (capsule_pages)

def trim_up(str):
    return str.replace(' ', '').replace(chr(194), '').replace(chr(160), '').upper()

def check_warnings(items_data, capsule_pages):
    for item in items_data:

        warn = ({"ЦЕНА": "Y",  # нулевая или нечисленная цена
                 "СТОРОННИЕ_ССЫЛКИ": "Y",  # сторонние ссылки
                 "НЕТ_РАЗМЕРОВ": "Y",  # нет размеров
                 "МОДЕЛЬ_КАТАЛОГ": "Y",  # различаются наименования в каталоге и карточке
                 "БРЕНД_КАТАЛОГ": "Y",  # различаются бренды в каталоге и карточке
                 "МОДЕЛЬ_TITLE": "Y",  # TITLE отличается от наименования
                 "ОТСУТСТВУЕТ_КРОСС_ССЫЛКА": "Y", #ссылка на эту страницу отсутствует на страницах других цветов
                 "МОДЕЛЬ_ВЕРСТКА": "Y", #различные название в верстке для десктопа и мобилы
                 "БРЕНД_ВЕРСТКА": "Y", #различные бренды в верстке для десктопа и мобилы
                 "ОПИСАНИЕ_ВЕРСТКА": "Y"
                 })

        try:
            n = int(item["itemPrice"])
            warn.pop("ЦЕНА")
        except:
            pass

        if len(item["foreignLinks"]) == 0:
            warn.pop("СТОРОННИЕ_ССЫЛКИ")

        if trim_up(item["itemNameD"]) == trim_up(item["pageTitle"]):
            warn.pop("МОДЕЛЬ_TITLE")

        if trim_up(item["itemNameD"]) == trim_up(item["catName"]):
            warn.pop("МОДЕЛЬ_КАТАЛОГ")

        if trim_up(item["itemBrandD"]) == trim_up(item["catBrand"]):
            warn.pop("БРЕНД_КАТАЛОГ")

        if (len(item["itemSizesStr"].strip()) != 0) or (item["chapter"] in ['bags', 'accessories']):
            warn.pop("НЕТ_РАЗМЕРОВ")

        if trim_up(item["itemNameD"]) == trim_up(item["itemNameM"]):
            warn.pop("МОДЕЛЬ_ВЕРСТКА")

        if trim_up(item["itemBrandD"]) == trim_up(item["itemBrandM"]):
            warn.pop("БРЕНД_ВЕРСТКА")

        if trim_up(item["itemDescD"]) == trim_up(item["itemDescM"]):
            warn.pop("ОПИСАНИЕ_ВЕРСТКА")


        selfURL = item["itemURL"]

        if item["otherColorsHref"] == []:
            warn.pop("ОТСУТСТВУЕТ_КРОСС_ССЫЛКА")
        else:
            for link in item["otherColorsHref"]:
                flag = False
                for d in items_data:
                    if d["itemURL"] == MainDomain + link:
                        for checklink in d["otherColorsHref"]:
                            if MainDomain + checklink == selfURL:
                                flag = True
                                break
                if not flag:
                    item["missedLinks"].append(link)

            if len(item["missedLinks"]) == 0:
                warn.pop("ОТСУТСТВУЕТ_КРОСС_ССЫЛКА")

        '''
        if item["capsuleLinks"] == []:
            warn.pop("КАПСУЛА_КРОСС_ССЫЛКА")
        else:
            for link in item["capsuleLinks"]:
                #print (f'link={link}')
                flag = False
                for d in items_data:
                    #print(f'itemURL={d["itemURL"]}')
                    #print(f'capsuleLinks={d["capsuleLinks"]}')
                    if d["itemURL"] == MainDomain + link:
                        #print ('in')
                        for chklink in d["capsuleLinks"]:
                            #print (f'chklink={chklink}')
                            if MainDomain + chklink == selfURL:
                                flag = True
                                break
                if not flag:
                    #print(f'not found={link}')
                    item["missedCapsuleLinks"].append(link)

            if len(item["missedCapsuleLinks"]) == 0:
                warn.pop("КАПСУЛА_КРОСС_ССЫЛКА")
        '''
        if len(warn.keys()) == 0:
            warn = ({"ОК": "Y"})
        item["warn"] = warn

    for capsule in capsule_pages:
        #print (f'capsule {capsule["code"]}')
        warn = ({"КАПСУЛА_КОЛИЧЕСТВО": "Y",  #неверное количество ссылок в капсуле
                 "КАПСУЛА_КРОСС_ССЫЛКА": "Y" #отсутствует кросс-ссылка для товаров из одной капсулы
                 })
        if len(capsule["items"]) == capsule["cnt"]:
            warn.pop("КАПСУЛА_КОЛИЧЕСТВО")

        for link in capsule["items"]:
            #print (f'link={link}')
            flag = False
            for d in items_data:
                #print(f'itemURL={d["itemURL"]}')

                if d["itemURL"] == MainDomain + link:
                    if d["capsuleLinks"] == capsule["items"]:
                        flag = True
                        #print ("OK")
                        break
                    #else:
                        #print (f'd = {d["capsuleLinks"]}')
                        #print (f'cap = {capsule["items"]}')
            if not flag:
                # print(f'not found={link}')
                capsule["missedItems"].append(link)

        if len(capsule["missedItems"]) == 0:
            warn.pop("КАПСУЛА_КРОСС_ССЫЛКА")

        if len(warn.keys()) == 0:
            warn = ({"ОК": "Y"})
        capsule["warn"] = warn

def parse_items_pages (data):
    idx = 0
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    browser = webdriver.Chrome(options=options)

    for item_page in data:
        get_page_data(item_page, browser, False)
        idx += 1
        if idx%20 == 0:
            print (f'{lib.cur_time()} parsed {idx}')

    browser.quit()

def get_pages_links(url, browser):
    print (url)

    browser.get(url + TAG)
    time.sleep(3)
    response = browser.page_source

    soup = BeautifulSoup(response, "lxml")

    pageBlocks = soup.find('div', class_ = 't786').find_all('div', class_='js-product t-store__card t-store__stretch-col t-store__stretch-col_33 t-align_left t-item') #js-product t-store__card t-store__stretch-col t-store__stretch-col_33 t-align_left t-item

    for pageBlock in pageBlocks:
        pages = pageBlock.find_all('a')
        #print(pages)
        for page in pages:
            p = page.get('href')
            try:
                Name = page.find('div', class_='js-store-prod-name js-product-name t-store__card__title t-name t-name_md').text.strip()
            except:
                Name = ''
            try:
                Brand = page.find('div', class_='js-store-prod-descr t-store__card__descr t-descr t-descr_xxs').text.replace('#', '').strip()
            except:
                Brand = ''

            if p[0] == '/':
                pageURL = MainDomain + p
                items_pages.append({"URL": pageURL, "Name": Name, "Brand": Brand, "Chapter": url})
    #print (len(items_pages))

def main():

    print(f"{lib.cur_time()} START")

    start_time = time.time()

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')

    ChapterList = ['/shoes', '/accessories', '/clothes', '/bags']
    #ChapterList = ['/bags']
    for Chapter in ChapterList:
        browser = webdriver.Chrome(options=options)
        browser.delete_all_cookies()
        get_pages_links(MainDomain + Chapter, browser)

    print(f"{lib.cur_time()} Parsing {len(items_pages)} pages")
    parse_items_pages(items_pages)

    print(f"{lib.cur_time()} Parsing capsules")
    capsule_pages = get_capsules()

    print(f"{lib.cur_time()} Checking warnings")
    check_warnings(items_data, capsule_pages)

    print(f"{lib.cur_time()} Saving XLS")
    file_name = save_xls ('induo', items_data, capsule_pages)

    finish_time = time.time() - start_time
    print(f"{lib.cur_time()} FINISH Total time: {finish_time}")
    return file_name

if __name__ == "__main__":

    bot = telebot.TeleBot(BOT_TOKEN)

    #get_capsules()

    #browser = webdriver.Chrome()
    #get_page_data({'URL': 'https://induo.store/checkered_shirt', 'Chapter': 'bags', 'Name': '', 'Brand': ''}, browser, True)



    # Функция, обрабатывающая команду /start
    @bot.message_handler(commands=["start"])
    def start(m, res=False):
        bot.send_message(m.chat.id, '!!)')

    '''
    # Получение сообщений от юзера
    @bot.message_handler(content_types=["text"])
    def handle_text(message):
        bot.send_message(message.chat.id, f'Сейчас посчитаю...')
        numPages = get_pages_links()
        bot.send_message(message.chat.id, f'Количество карточек: {numPages}')
    '''

    @bot.message_handler(commands=["file"])
    def handle_file(message):
        bot.send_message(message.chat.id, f'Формирую файл')
        file_name = main()
        with open(f"../output/{file_name}", "rb") as misc:
            f = misc.read()
        bot.send_document(chat_id = message.chat.id, document = f, visible_file_name =file_name)

    # Запускаем бота
    bot.polling(none_stop=True, interval=0)